import os
import io
import json
from datetime import date

import pandas as pd
from fastapi import APIRouter, File, UploadFile, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..deps import get_db, get_current_user
from .. import models

router = APIRouter(prefix="/api/v1/uploads", tags=["uploads"])

UPLOAD_DIR = os.path.join(os.getcwd(), "uploads_temp")
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Get or create a Product (Tower) by name
# ─────────────────────────────────────────────────────────────────────────────
def get_or_create_product(db: Session, name: str, organization_id: int, fields: list) -> models.Product:
    product = db.query(models.Product).filter(
        models.Product.name == name,
        models.Product.organization_id == organization_id
    ).first()

    if not product:
        product = models.Product(
            name=name,
            organization_id=organization_id,
            description="",
            input_fields=[],
            output_fields=fields,  # treat all metric columns as output fields
            sector="Telecom"
        )
        db.add(product)
        db.flush()

    return product


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL UPLOAD — Flat Single-Sheet Format
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Accept an Excel file with a single 'Data' sheet (or the first sheet).
    Required columns: Date, Tower
    All remaining numeric columns are treated as data fields.

    Raises a 400 error if Date or Tower columns are missing.
    """
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        sheet = df.fillna("")

        # Check required columns
        field_columns = list(sheet.columns)
        if "Date" not in field_columns or "Tower" not in field_columns:
            raise HTTPException(status_code=400, detail="Missing required columns: 'Date' and 'Tower'")

        # Clean up
        sheet = sheet.dropna(how="all")
        sheet.columns = [str(c).strip() for c in sheet.columns]

        # Identify data field columns (everything except Date and Tower)
        field_columns = [c for c in sheet.columns if c not in ("Date", "Tower")]
        if not field_columns:
            raise HTTPException(status_code=400, detail="No data field columns found (e.g. Grid_kWh, Revenue_PKR).")

        records_created = 0
        records_updated = 0
        errors = []

        # If data is flat (legacy), we'll just treat it as a single tenant named "Main"
        is_multi_tenant = "Tenant" in field_columns
        data_fields = [c for c in field_columns if c != "Tenant"]

        # Group by Date and Tower
        groups = sheet.groupby(['Date', 'Tower'])

        for (date_val, tower_val), group in groups:
            try:
                date_val = str(date_val).strip()
                tower_val = str(tower_val).strip()

                if not date_val or tower_val in ("", "nan"):
                    continue

                # Get/create the product (tower)
                product = get_or_create_product(
                    db, tower_val, user.organization_id, data_fields
                )

                tenants_list = []
                total_tower_revenue = 0

                for idx, row in group.iterrows():
                    tenant_name = str(row.get('Tenant', "Main Tenant")) if is_multi_tenant else "Main Tenant"
                    
                    # Extract inputs
                    inputs = {}
                    for field in data_fields:
                        if field not in ['PricePerKilowatt', 'TotalRevenue', 'TotalKilowattsProduced', 'TotalKilowattsSold']:
                            val = row.get(field)
                            try: inputs[field] = float(val) if pd.notna(val) else 0.0
                            except: inputs[field] = 0.0
                    
                    # Output assumptions
                    produced_val = row.get('TotalKilowattsProduced', row.get('Grid_kWh', 0))
                    try: produced = float(produced_val) if pd.notna(produced_val) else 0.0
                    except: produced = 0.0

                    price_val = row.get('PricePerKilowatt', 22)
                    try: price = float(price_val) if pd.notna(price_val) else 22.0
                    except: price = 22.0

                    revenue = produced * price
                    if "Revenue_PKR" in row:
                        try: revenue = float(row.get("Revenue_PKR")) if pd.notna(row.get("Revenue_PKR")) else revenue
                        except: pass

                    inputs['totalKilowattsProduced'] = produced

                    tenants_list.append({
                        "name": tenant_name,
                        "inputs": inputs,
                        "pricePerKilowatt": price,
                        "outputs": {
                            "kilowattsSold": produced,
                            "totalRevenue": revenue
                        }
                    })
                    total_tower_revenue += revenue

                # Check if multi-tenant record exists for this Tower and Date
                existing_record = db.query(models.ProductDataRecord).filter(
                    models.ProductDataRecord.product_id == product.id,
                    models.ProductDataRecord.month == date_val,  # Repurposing 'month' db col for date
                    models.ProductDataRecord.organization_id == user.organization_id
                ).first()

                if existing_record:
                    existing_record.data = {
                        "parameters": {
                            "date": date_val,
                            "towerName": product.name,
                            "location": product.description
                        },
                        "tenants": tenants_list,
                        "totalTowerRevenue": total_tower_revenue
                    }
                    records_updated += 1
                else:
                    record = models.ProductDataRecord(
                        organization_id=user.organization_id,
                        product_id=product.id,
                        month=date_val,
                        data={
                            "parameters": {
                                "date": date_val,
                                "towerName": product.name,
                                "location": product.description
                            },
                            "tenants": tenants_list,
                            "totalTowerRevenue": total_tower_revenue
                        }
                    )
                    db.add(record)
                    records_created += 1

            except Exception as e:
                errors.append(f"Error processing {tower_val} / {date_val}: {str(e)}")

        db.commit()

        return {
            "success": True,
            "message": f"Upload complete. {records_created} records created, {records_updated} updated.",
            "records_created": records_created,
            "records_updated": records_updated,
            "errors": errors
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = f"System Error: {str(e)}"
        print(f"UPLOAD ERROR: {error_detail}")
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=error_detail) # Return 400 instead of 500 to keep user logged in if possible


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE DOWNLOAD — Multi-Tenant Format
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/template")
def download_excel_template():
    """
    Download the multi-tenant Excel template matching the new required data format.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row
    headers = [
        "Date", "Tower", "Tenant", "TotalKilowattsProduced", "PricePerKilowatt", 
        "FuelExpense", "HRExpense", "OperationExpense", "MaintenanceExpense", 
        "BatteryExpense", "SolarGridExpense", "DieselFuelExpense"
    ]
    ws.append(headers)

    # Style the header
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Example data rows
    example_rows = [
        ["2026-04-17", "TLN-KHI-01", "Tenant A", 5200, 22, 10000, 5000, 2000, 1000, 0, 0, 6500],
        ["2026-04-17", "TLN-KHI-01", "Tenant B", 4800, 25, 8000,  5000, 2500,  800, 0, 0, 5500],
        ["2026-04-18", "TLN-LHR-02", "Tenant A", 5500, 20, 12000, 5000, 2000, 1000, 0, 0, 7000],
        ["2026-04-19", "TLN-KHI-01", "Tenant A", 5300, 22, 10500, 5000, 2000, 1000, 0, 0, 6600],
        ["2026-04-19", "TLN-KHI-01", "Tenant B", 4900, 25, 8200,  5000, 2500,  800, 0, 0, 5600],
    ]

    for row in example_rows:
        ws.append(row)

    # Style the data rows
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center")

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Instruction sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Productix Multi-Tenant Data Upload Template"],
        [""],
        ["COLUMNS:"],
        ["Date", "The specific date for this data (e.g. 2026-04-17) - REQUIRED"],
        ["Tower", "The tower/site name (acts as the Product ID) - REQUIRED"],
        ["Tenant", "The name of the tenant occupying the tower - REQUIRED"],
        ["TotalKilowattsProduced", "KW produced for this tenant"],
        ["PricePerKilowatt", "Price charged per KW"],
        ["*Expense", "Various input expense fields for the tenant"],
        [""],
        ["RULES:"],
        ["- Add multiple rows with the same Date and Tower, but different Tenants."],
        ["- All metric columns form the inputs for that specific Tenant."],
        ["- If a Tower does not exist, it will be created automatically."],
        ["- If records for the same Tower+Date exist, they will be OVERWRITTEN with the new Tenants."],
    ]
    for row in instructions:
        ws2.append(row)

    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 65

    # Write to stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="productix_multi_tenant_template.xlsx"'}
    )

